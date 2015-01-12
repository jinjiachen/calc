#coding=cp936

from openpyxl import Workbook
from openpyxl import load_workbook
import math

def exc(i,num,name,mat,pai,x1,x2,x3,amount):
    wb=load_workbook('D:\samples.xlsx')
    ws=wb.active
    ws.cell(row=i,column=2).value=num
    ws.cell(row=i,column=3).value=name
    ws.cell(row=i,column=6).value=pai
    ws.cell(row=i,column=4).value=amount
    if mat.encode('cp936')=='方钢板':
        mat=u'钢板'
        ws.cell(row=i,column=5).value=mat
        ws.cell(row=i,column=7).value=x1+u'×'+x2+',t='+x3
        m1=plate(x1,x2,x3,amount)
        ws.cell(row=i,column=13).value=m1
        x1=str(float(x1)+10)
        x2=str(float(x2)+10)
        ws.cell(row=i,column=8).value=x1+u'×'+x2+',t='+x3
        m2=plate(x1,x2,x3,amount)
        ws.cell(row=i,column=14).value=m2
        ws.cell(row=i,column=16).value=93
        ws.cell(row=i,column=15).value=m2/0.93
    elif mat.encode('cp936')=='圆钢板':
        mat=u'钢板'
        ws.cell(row=i,column=5).value=mat
        ws.cell(row=i,column=7).value=u'φ'+x1+',t='+x2
        m1=circle(x1,x2,amount)
        ws.cell(row=i,column=13).value=m1
        x1=str(float(x1)+10)
        ws.cell(row=i,column=8).value=u'φ'+x1+',t='+x2
        m2=circle(x1,x2,amount)
        ws.cell(row=i,column=14).value=m2
        ws.cell(row=i,column=16).value=72
        ws.cell(row=i,column=15).value=m2/0.72
    elif mat.encode('cp936')=='钢管':
        ws.cell(row=i,column=5).value=mat
        ws.cell(row=i,column=7).value=u'φ'+x1+u'×'+x2+',L='+x3
        m1=pipe(x1,x2,x3,amount)
        ws.cell(row=i,column=13).value=m1
        x3=str(float(x3)+10)
        ws.cell(row=i,column=8).value=u'φ'+x1+u'×'+x2+',L='+x3
        m2=pipe(x1,x2,x3,amount)
        ws.cell(row=i,column=14).value=m2
        ws.cell(row=i,column=16).value=88
        ws.cell(row=i,column=15).value=m2/0.88
    elif mat.encode('cp936')=='圆钢':
        ws.cell(row=i,column=5).value=mat
        ws.cell(row=i,column=7).value=u'φ'+x1+u'×'+x2
        m1=cylinder(x1,x2,amount)
        ws.cell(row=i,column=13).value=m1
        x1=str(float(x1)+5)
        x2=str(float(x2)+10)
        ws.cell(row=i,column=8).value=u'φ'+x1+u'×'+x2
        m2=cylinder(x1,x2,amount)
        ws.cell(row=i,column=14).value=m2
        ws.cell(row=i,column=16).value=92
        ws.cell(row=i,column=15).value=m2/0.92

    wb.save('D:\samples.xlsx')



def plate(x1,x2,x3,amount):
    x1=float(x1)
    x2=float(x2)
    x3=float(x3)
    m=x1*x2*x3/1000**3*7850*amount
    return(m)

def pipe(x1,x2,x3,amount):
    Do=float(x1)
    t=float(x2)
    L=float(x3)
    Di=Do-2*t
    m=math.pi/4*(Do**2-Di**2)*L/1000**3*7850*amount
    return(m)

def cylinder(x1,x2,amount):
    Do=float(x1)
    L=float(x2)
    m=math.pi/4*Do**2*L/1000**3*7850*amount
    return(m)
        
def circle(x1,x2,amount):
    Do=float(x1)
    t=float(x2)
    m=math.pi/4*Do**2*t/1000**3*7850*amount
    return(m)



    
